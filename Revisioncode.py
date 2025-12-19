import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import tempfile
import random
import smtplib
import time
from email.mime.text import MIMEText

# =========================
# PAGE CONFIG
# =========================
st.set_page_config(
    page_title="R0 vs R1 Tag Comparison",
    page_icon="📊",
    layout="wide"
)

# =========================
# OTP CONFIG (USE APP PASSWORD)
# =========================
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
SENDER_EMAIL = "yourgmail@gmail.com"
SENDER_PASSWORD = "your_gmail_app_password"
OTP_EXPIRY_SECONDS = 300  # 5 minutes

# =========================
# SESSION STATE INIT
# =========================
defaults = {
    "authenticated": False,
    "otp": None,
    "otp_time": None,
    "email": None,
    "run": False,
    "completed": False
}
for k, v in defaults.items():
    st.session_state.setdefault(k, v)

# =========================
# OTP FUNCTIONS
# =========================
def generate_otp():
    return str(random.randint(100000, 999999))

def send_otp(email, otp):
    msg = MIMEText(f"Your login OTP is {otp}\n\nValid for 5 minutes.")
    msg["Subject"] = "Streamlit Login OTP"
    msg["From"] = SENDER_EMAIL
    msg["To"] = email

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.send_message(msg)

# =========================
# LOGIN GATE
# =========================
if not st.session_state.authenticated:
    st.title("🔐 Secure Login")

    email = st.text_input("Enter your email")

    if st.button("Send OTP"):
        if not email:
            st.error("Email is required")
        else:
            otp = generate_otp()
            send_otp(email, otp)
            st.session_state.otp = otp
            st.session_state.otp_time = time.time()
            st.session_state.email = email
            st.success("OTP sent to your email")

    if st.session_state.otp:
        user_otp = st.text_input("Enter OTP", max_chars=6)

        if st.button("Verify OTP"):
            if time.time() - st.session_state.otp_time > OTP_EXPIRY_SECONDS:
                st.error("OTP expired. Request a new one.")
                st.session_state.otp = None
            elif user_otp == st.session_state.otp:
                st.session_state.authenticated = True
                st.success("Login successful")
                st.rerun()
            else:
                st.error("Invalid OTP")

    st.stop()  # 🚫 Stop app until authenticated

# =========================
# MAIN APP (AFTER LOGIN)
# =========================
st.title("📊 R0 vs R1 Tag Comparison")
st.caption(f"Logged in as {st.session_state.email}")

# =========================
# FILE UPLOAD
# =========================
col1, col2 = st.columns(2)

with col1:
    r0_file = st.file_uploader("Upload R0.xlsx", type=["xlsx"])

with col2:
    r1_file = st.file_uploader("Upload R1.xlsx", type=["xlsx"])

# =========================
# RUN BUTTON
# =========================
run_disabled = not (r0_file and r1_file)

if st.button("🚀 Run Comparison", disabled=run_disabled):
    st.session_state.run = True
    st.session_state.completed = False

# =========================
# STATUS + PROGRESS
# =========================
status_box = st.empty()
progress_bar = st.empty()

if not st.session_state.run:
    status_box.info("Upload both files and click Run Comparison")

if st.session_state.run and not st.session_state.completed:
    status_box.warning("Processing...")
    progress = progress_bar.progress(0)

    progress.progress(10)
    r0_df = pd.read_excel(r0_file, dtype=str).fillna("")
    r1_df = pd.read_excel(r1_file, dtype=str).fillna("")

    if "Tag" not in r0_df.columns or "Tag" not in r1_df.columns:
        st.error("Both files must contain a 'Tag' column.")
        st.stop()

    progress.progress(30)
    r0_df = r0_df.drop_duplicates(subset="Tag").set_index("Tag")
    r1_df = r1_df.drop_duplicates(subset="Tag").set_index("Tag")

    r0_columns = list(r0_df.columns)
    all_columns = sorted(
        set(r0_df.columns).union(set(r1_df.columns)),
        key=lambda x: r0_columns.index(x) if x in r0_columns else float("inf")
    )

    all_tags = sorted(set(r0_df.index).union(set(r1_df.index)))

    progress.progress(60)
    comparison_rows = []

    for tag in all_tags:
        if tag not in r0_df.index:
            row = {"Tag": tag, "Change_Type": "Added in R1"}
            row.update({col: r1_df.loc[tag].get(col, "") for col in all_columns})
            row["Change_Summary"] = ""
            comparison_rows.append(row)

        elif tag not in r1_df.index:
            row = {"Tag": tag, "Change_Type": "Removed in R1"}
            row.update({col: r0_df.loc[tag].get(col, "") for col in all_columns})
            row["Change_Summary"] = ""
            comparison_rows.append(row)

        else:
            row_r0 = r0_df.loc[tag]
            row_r1 = r1_df.loc[tag]
            row_data = {"Tag": tag}
            summary = []
            changed = False

            for col in all_columns:
                v0 = row_r0.get(col, "")
                v1 = row_r1.get(col, "")
                if str(v0).strip() != str(v1).strip():
                    row_data[col] = f"{v0} → {v1}"
                    summary.append(f"{col}: {v0} → {v1}")
                    changed = True
                else:
                    row_data[col] = v1

            row_data["Change_Type"] = "Modified" if changed else "No Change"
            row_data["Change_Summary"] = " | ".join(summary)
            comparison_rows.append(row_data)

    progress.progress(85)
    comparison_df = pd.DataFrame(comparison_rows)
    final_columns = ["Tag", "Change_Type"] + all_columns + ["Change_Summary"]
    comparison_df = comparison_df[final_columns]

    progress.progress(100)
    progress_bar.empty()
    status_box.success("Comparison completed")
    st.session_state.completed = True

# =========================
# RESULTS + EXPORT
# =========================
if st.session_state.completed:
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Tags in R0", len(r0_df))
    c2.metric("Tags in R1", len(r1_df))
    c3.metric("Added", len(set(r1_df.index) - set(r0_df.index)))
    c4.metric("Removed", len(set(r0_df.index) - set(r1_df.index)))

    st.dataframe(comparison_df, use_container_width=True, height=500)

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for r_idx, row in enumerate(dataframe_to_rows(comparison_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if isinstance(value, str) and "→" in value:
                cell.fill = highlight

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        output_path = tmp.name

    st.download_button(
        "⬇️ Download Comparison Excel",
        data=open(output_path, "rb"),
        file_name=f"Vimal_Comparison_{datetime.now().strftime('%d_%m_%Y_%H_%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =========================
# LOGOUT
# =========================
if st.sidebar.button("Logout"):
    st.session_state.clear()
    st.rerun()
